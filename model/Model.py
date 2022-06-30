import pandas as pd
import json
import itertools
from openpyxl import load_workbook


class Model:
    # Constructor for Model
    def __init__(self, turn):
        self._paramsDict = dict()  # Not used yet
        self._paramsList = list()  # Used to store values from excel file
        self._columNames = ["WACC",
                            "CO2-pris",
                            "CO2-utsläpp per MWh",
                            "CO2-skatt per MWh",
                            "Pris ren energi",
                            "Pris CO2-energi",
                            "de facto pris CO2-energi",
                            "Differens ren och smutsig energi",
                            "Energikonsumtion",
                            "Andel CO2-energi",
                            "Start EBIT",
                            "CO2 utsläpp industri"]  # Used to store the column names from excel file
        self.initValues("excelData.xlsx")
        self.decisionsData = self.getDecisionsForTurnX("beslut.json", turn)
        self.calculateParamsForAllDecisionCombinations()

    def initValues(self, excelPath):
        """Initializes the class Dict and List with the params from a excel file
        """
        # Load in excel file data to excel_in
        excel_in = pd.read_excel(r''+excelPath)
        # Select first row in excel file
        originalParams = excel_in.iloc[0]

        # test
        wb = load_workbook(filename=excelPath)
        sheet_names = wb.sheetnames
        name = sheet_names[0]
        sheet_ranges = wb[name]
        df = pd.DataFrame(sheet_ranges.values)
        print(df)
        originalParams = df.iloc[1]

        # Iterate through the params and add them to the dict and list
        for index, paramValue in enumerate(originalParams):
            paramName = self._columNames[index]
            self._paramsDict[paramName] = paramValue
            self._paramsList.append(paramValue)

    def getDecisionsForTurnX(self, decisionJson, turn):
        """Returns a dict with the decisions for turn X given a decision json file"""
        with open(decisionJson) as json_file:
            data = json.load(json_file)
            return data[turn]

    def calculateParamsForAllDecisionCombinations(self):
        """Calculates the new params for all decision combinations and stores them to excel"""

        # Create a list that containts lists of all sectors decisions e.g. [["A", "B"], ["C", "D"]]
        listOfDecisionLists = list()
        for sector in self.decisionsData:
            listOfDecisionLists.append(self.decisionsData[sector])

        # Store the new lists of params in this list
        newParams = list()

        # Creates a list of tuples of all possible combinations
        listOfCombinations = list(itertools.product(*listOfDecisionLists))

        # Iterate through all the tuples of combinations and calculate the new params for that given tuple of decisions
        for uniqueCombo in listOfCombinations:
            params = self.calculateNewParamsValuesFromDecisionCombinations(
                uniqueCombo)
            newParams.append(params)

        self.saveNewParamsToExcel(newParams)

    def calculateNewParamsValuesFromDecisionCombinations(self, comboOfDecisions):
        """Calculates the new params for a given decision combination and returns a list of the new params"""
        copyOfParams = self._paramsList.copy()
        for decision in comboOfDecisions:
            self.doDecisionCalculation(decision, copyOfParams)

        # Innan vi är färdiga och skickar tillbaka de nya värderna
        # Gör om beräkningar på parametrar som beror på andra parametrar
        # så att de uppdateras om en underliggande parameter har ändras i doDecisionCalculation
        self.recalcDependenParams(copyOfParams)
        return copyOfParams

    def doDecisionCalculation(self, decision, params):
        """Go through a given decisions effect and calculate how the given params should change"""

        effect = decision["Effekt"]
        listOfParamsToChange = effect["Parametrar"]
        listOfTypeOfChanges = effect["TypeOfChange"]
        listOfTheChangeValues = effect["changeValue"]

        for index, param in enumerate(listOfParamsToChange):
            if(listOfTypeOfChanges[index] == "addition"):
                params[self.paramNameToIndex(
                    param)] += listOfTheChangeValues[index]
            elif(listOfTypeOfChanges[index] == "factorMultiplication"):
                params[self.paramNameToIndex(
                    param)] *= listOfTheChangeValues[index]

    def paramNameToIndex(self, paramName):
        """Returns the index of a param in the list"""
        convertTable = {
            "WACC": 0,
            "CO2-pris": 1,
            "CO2-utsläpp per MWh": 2,
            "CO2-skatt per MWh": 3,
            "Pris ren energi": 4,
            "Pris CO2-energi": 5,
            "de facto pris CO2-energi": 6,
            "Differens ren och smutsig energi": 7,
            "Energikonsumtion": 8,
            "Andel CO2-energi": 9,
            "Start EBIT": 10,
            "CO2 utsläpp industri": 11,
        }
        return convertTable[paramName]

    def indexToParamName(self, index):
        """Returns the name of a param given the index"""
        convertTable = ["WACC",
                        "CO2-pris",
                        "CO2-utsläpp per MWh",
                        "CO2-skatt per MWh",
                        "Pris ren energi",
                        "Pris CO2-energi",
                        "de facto pris CO2-energi",
                        "Differens ren och smutsig energi",
                        "Energikonsumtion",
                        "Andel CO2-energi",
                        "Start EBIT",
                        "CO2 utsläpp industri"]

        return convertTable[index]

    def saveNewParamsToExcel(self, paramsList):
        """Saves a list of paramlists to excel"""
        # print(paramsList)
        paramsList.insert(0, self._columNames)
        tempDataFrame = pd.DataFrame(paramsList)
        tempDataFrame.to_excel("test.xlsx")
        return None

    def recalcDependenParams(self, paramList):
        """Takes a list of params and recalcs the dependent ones"""
        for index, param in enumerate(paramList):
            name = self.indexToParamName(index)
            print(name)
            if(name == "CO2-skatt per MWh"):
                paramList[index] = paramList[self.paramNameToIndex(
                    "CO2-pris")] * paramList[self.paramNameToIndex("CO2-utsläpp per MWh")]
            elif(name == "de facto pris CO2-energi"):
                paramList[index] = paramList[self.paramNameToIndex(
                    "CO2-skatt per MWh")] + paramList[self.paramNameToIndex("Pris CO2-energi")]
            elif(name == "Differens ren och smutsig energi"):
                paramList[index] = paramList[self.paramNameToIndex(
                    "de facto pris CO2-energi")] - paramList[self.paramNameToIndex("Pris ren energi")]
            elif(name == "CO2 utsläpp industri"):
                paramList[index] = paramList[self.paramNameToIndex("CO2-utsläpp per MWh")] * paramList[self.paramNameToIndex(
                    "Energikonsumtion")] * paramList[self.paramNameToIndex("Andel CO2-energi")]


def main():
    m = Model(0)

if(__name__ == "__main__"):
    main()
